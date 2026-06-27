"""Report and log generation service."""
from __future__ import annotations
import os
import time
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.schemas import ValidationResult, ConversionResult


def _now_fix() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def generate_migration_report(
    filename: str,
    result: ConversionResult,
    output_path: str,
) -> None:
    lines = [
        "=" * 60,
        "  SHOPIFY → WOOCOMMERCE MIGRATION REPORT",
        "=" * 60,
        "",
        f"Generated:        {_now_fix()}",
        f"Source file:      {filename}",
        "",
        "─" * 60,
        "  CONVERSION SUMMARY",
        "─" * 60,
        f"  Products converted:   {result.products_converted}",
        f"  Products failed:      {result.products_failed}",
        f"  Variants converted:   {result.variants_converted}",
        f"  Categories mapped:    {result.categories_mapped}",
        f"  Tags preserved:       {result.tags_preserved}",
        f"  Images mapped:        {result.images_mapped}",
        f"  Warnings:             {result.warnings}",
        f"  Errors:               {result.errors}",
        f"  Execution time:       {result.execution_time_seconds:.2f}s",
        "",
        "─" * 60,
        "  OUTPUT FILES",
        "─" * 60,
    ]
    for f in result.output_files:
        lines.append(f"  • {f}")
    lines += [
        "",
        "─" * 60,
        "  LOG (last 100 entries)",
        "─" * 60,
    ]
    lines += result.log_lines[-100:]
    lines += ["", "=" * 60]

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_validation_xlsx(
    filename: str,
    result: ValidationResult,
    output_path: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"  # type: ignore[union-attr]

    # Colour palette
    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
    ERR_FILL = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")

    PASS_FONT = Font(color="276221", bold=False)
    WARN_FONT = Font(color="9C6500", bold=False)
    ERR_FONT = Font(color="9C0006", bold=False)
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:F1")  # type: ignore[union-attr]
    title_cell = ws["A1"]  # type: ignore[index]
    title_cell.value = f"Validation Report — {filename}"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    title_cell.alignment = Alignment(horizontal="left")

    ws["A2"] = f"Generated: {_now_fix()}"  # type: ignore[index]
    ws["A2"].font = Font(italic=True, color="595959")  # type: ignore[union-attr]

    # Summary row
    ws["A3"] = f"PASS: {result.pass_count}   WARNING: {result.warning_count}   ERROR: {result.error_count}"  # type: ignore[index]

    # Header row
    headers = ["Level", "Check", "Message", "Count", "Can Convert?", "Details (first 10)"]
    ws.append([])  # blank row
    ws.append(headers)  # type: ignore[union-attr]

    header_row_idx = ws.max_row  # type: ignore[union-attr]
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)  # type: ignore[union-attr]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Data rows
    for issue in result.issues:
        details_str = "; ".join(issue.details[:10])
        can_convert = "" if issue.level != "error" else "Blocks conversion"
        ws.append([issue.level.upper(), issue.check, issue.message, issue.count, can_convert, details_str])  # type: ignore[union-attr]

        row_idx = ws.max_row  # type: ignore[union-attr]
        if issue.level == "pass":
            fill, font_ = PASS_FILL, PASS_FONT
        elif issue.level == "warning":
            fill, font_ = WARN_FILL, WARN_FONT
        else:
            fill, font_ = ERR_FILL, ERR_FONT

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore[union-attr]
            if col_idx == 1:
                cell.fill = fill
                cell.font = font_
                cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if col_idx in (3, 6):
                cell.alignment = Alignment(wrap_text=True)

    # Column widths
    col_widths = [12, 30, 60, 10, 18, 80]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width  # type: ignore[union-attr]

    wb.save(output_path)


def generate_conversion_log(
    log_lines: list[str],
    output_path: str,
) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
